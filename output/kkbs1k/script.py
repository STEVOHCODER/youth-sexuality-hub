import sys
import traceback

try:
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, PatternFill, Alignment

    print("⚙️ Initializing Stochastic Brownian Motion Simulation...")

    # 1. MATHEMATICAL MODELING: Generate Data
    # Simulating random particle displacement governed by diffusion
    np.random.seed(42)
    time_steps = 100
    particles = 5
    
    # Generate standard normal steps and compute cumulative sum (Random Walk)
    steps = np.random.normal(loc=0, scale=1.0, size=(time_steps, particles))
    positions = np.cumsum(steps, axis=0)
    
    columns = [f"Particle_{i+1}" for i in range(particles)]
    df = pd.DataFrame(positions, columns=columns)
    df.insert(0, 'Time_Step', np.arange(1, time_steps + 1))
    
    # Calculate statistical metrics
    df['Mean_Displacement'] = df[columns].mean(axis=1)
    df['Variance'] = df[columns].var(axis=1)

    # 2. DATA VISUALIZATION (Safe for Preview Environment)
    plt.figure(figsize=(10, 6))
    for col in columns:
        plt.plot(df['Time_Step'], df[col], alpha=0.8, linewidth=1.5, label=col)
    
    plt.plot(df['Time_Step'], df['Mean_Displacement'], color='black', 
             linewidth=3, linestyle='--', label='Mean Displacement')
    
    plt.title("Statistical Physics: Particle Brownian Motion", fontsize=14, weight='bold')
    plt.xlabel("Time Step ($t$)", fontsize=12)
    plt.ylabel("Displacement ($x$)", fontsize=12)
    plt.legend(loc='upper left')
    plt.grid(True, linestyle=':', alpha=0.7)
    
    print("📊 Rendering plot in preview window...")
    plt.show()

    # 3. EXCEL FILE GENERATION
    print("📝 Assembling the Excel Workbook...")
    wb = Workbook()
    
    # --- Sheet 1: Simulation Data ---
    ws_data = wb.active
    ws_data.title = "Simulation Data"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    # Styling the header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # --- Sheet 2: Summary Dashboard ---
    ws_summary = wb.create_sheet("Summary Dashboard")
    summary_data = [
        ["Metric", "Value"],
        ["Simulation Type", "Stochastic Brownian Motion"],
        ["Total Time Steps", time_steps],
        ["Number of Particles", particles],
        ["Final Empirical Variance", float(df['Variance'].iloc[-1])],
        ["Max Variance (Excel Formula)", "=MAX('Simulation Data'!H2:H101)"] 
    ]
    
    for row in summary_data:
        ws_summary.append(row)

    for cell in ws_summary[1]:
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Save cleanly to the local environment directory
    file_name = "Statistical_Simulation_Report.xlsx"
    wb.save(file_name)
    
    print(f"\n✅ SUCCESS! The file '{file_name}' has been successfully generated.")
    print("You can now download the Excel file using your environment's file/artifact explorer.")

except ImportError as e:
    print(f"\n❌ MISSING DEPENDENCY ERROR: {e}")
    print("Please ensure 'pandas', 'openpyxl', and 'matplotlib' are available in this environment.")
except Exception as e:
    print(f"\n❌ SCRIPT ERROR:")
    print(traceback.format_exc())
